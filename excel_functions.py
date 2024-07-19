# Excel functions file to read  and write Excel file f
# use for writing and reading excel
from openpyxl import load_workbook


class Excel_function:

    def __init__(self, filename, sheet_number):
        self.file = filename
        self.sheet = sheet_number

    # method to count total no. of rows in excel file
    def row_count(self):
        work_book = load_workbook(self.file)
        sheet = work_book[self.sheet]
        return sheet.max_row

    # method to count no. of columns in excel file
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column

    # read data from excel file
    def read_data(self, row_number, column_number):
        work_book = load_workbook(self.file)
        sheet = work_book[self.sheet]
        return sheet.cell(row=row_number, column=column_number).value

    # write data in excel file
    def write_data(self, row_number, column_number, data):
        work_book = load_workbook(self.file)
        sheet = work_book[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        work_book.save(self.file)
